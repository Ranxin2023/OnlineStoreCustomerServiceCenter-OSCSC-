from openpyxl import load_workbook

file_path = "Yanwen小包专线模板_V251224.xlsx"

wb = load_workbook(file_path, data_only=True)
ws = wb.active

print("====== 所有带下拉的数据列 ======\n")

for dv in ws.data_validations.dataValidation:
    print("📌 范围:", dv.sqref)
    print("类型:", dv.type)
    print("来源:", dv.formula1)

    # 提取列字母（比如 D、J、Z）
    cols = set()
    for cell_range in str(dv.sqref).split():
        col = ''.join(filter(str.isalpha, cell_range.split(":")[0]))
        cols.add(col)

    print("👉 涉及列:", list(cols))
    print("-" * 50)

print("====== 字段结构分析 ======\n")

# 先收集所有有验证的列
validation_cols = set()

for dv in ws.data_validations.dataValidation:
    for cell_range in str(dv.sqref).split():
        col = ''.join(filter(str.isalpha, cell_range.split(":")[0]))
        validation_cols.add(col)

# 遍历所有列
for col in ws.iter_cols(min_row=1, max_row=1):
    col_letter = col[0].column_letter
    header = col[0].value

    if col_letter in validation_cols:
        print(f"{col_letter} | {header} | ✅ 下拉字段")
    else:
        print(f"{col_letter} | {header} | 普通字段")


