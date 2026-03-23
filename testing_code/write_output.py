"""
write_output.py
修复：dataValidations 必须在 <hyperlinks>/<pageMargins> 之前，
      extLst 在 </worksheet> 之前。完全按原模板的正确顺序注入。
"""
from openpyxl import load_workbook
import warnings
import zipfile
import os
import re
import copy
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

SRC = "Yanwen小包专线模板_V251224.xlsx"
TMP = "_output_tmp.xlsx"
OUT = "output_test.xlsx"

def safe_copy_style(src, tgt):
    for attr in ("font", "fill", "border", "alignment"):
        try:
            v = getattr(src, attr)
            if v: 
                setattr(tgt, attr, copy.copy(v))
        except Exception: 
            pass
    try:
        if src.number_format:
            tgt.number_format = src.number_format
    except Exception: 
        pass

# ── 1. openpyxl 写数据+样式 ──
wb = load_workbook(SRC)
ws = wb.active
ws.data_validations.dataValidation.clear()

template_row = 2
for i in range(2, 12):
    for col in range(1, ws.max_column + 1):
        safe_copy_style(ws.cell(row=template_row, column=col),
                        ws.cell(row=i, column=col))

for i in range(2, 12):
    ws[f"A{i}"] = f"100000{i}"
    ws[f"C{i}"] = "北京燕文"
    ws[f"D{i}"] = "大陆DHL"
    ws[f"E{i}"] = "TestUser"
    ws[f"F{i}"] = "123456789"
    ws[f"J{i}"] = "西班牙"
    ws[f"Z{i}"] = "美元"
    ws[f"AA{i}"] = "否"

wb.save(TMP)
print("✅ Step1: 数据写入完成")

# ── 2. 从原模板提取两个验证块 ──
with zipfile.ZipFile(SRC) as z:
    orig_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

std_dv  = re.search(r'<dataValidations\b.*?</dataValidations>', orig_xml, re.DOTALL)
ext_lst = re.search(r'<extLst>.*?</extLst>', orig_xml, re.DOTALL)

std_dv_xml  = std_dv.group(0)  if std_dv  else ""
ext_lst_xml = ext_lst.group(0) if ext_lst else ""
print(f"✅ Step2: 标准验证={'找到' if std_dv_xml else '无'}, extLst={'找到' if ext_lst_xml else '无'}")

# ── 3. 按正确顺序注入 ──
with zipfile.ZipFile(TMP) as z:
    tmp_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

# OOXML 正确顺序：dataValidations 在 <hyperlinks 或 <pageMargins 之前
# extLst 在 </worksheet> 之前
# 先找插入 dataValidations 的锚点（取最早出现的）
anchor_tags = ['<hyperlinks', '<pageMargins', '<pageSetup', '</worksheet>']
anchor_pos = len(tmp_xml)
anchor_tag_used = '</worksheet>'
for tag in anchor_tags:
    idx = tmp_xml.find(tag)
    if idx != -1 and idx < anchor_pos:
        anchor_pos = idx
        anchor_tag_used = tag

print(f"  dataValidations 插入锚点: '{anchor_tag_used}' @ pos={anchor_pos}")

# 在锚点前插入标准验证
if std_dv_xml:
    tmp_xml = tmp_xml[:anchor_pos] + std_dv_xml + tmp_xml[anchor_pos:]

# extLst 插入到 </worksheet> 前
tmp_xml = tmp_xml.replace("</worksheet>", ext_lst_xml + "</worksheet>")

# ── 4. 重新打包 ──
new_zip = OUT + ".new"
with zipfile.ZipFile(TMP, 'r') as zin, \
     zipfile.ZipFile(new_zip, 'w', zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        if item.filename == "xl/worksheets/sheet1.xml":
            zout.writestr(item, tmp_xml.encode("utf-8"))
        else:
            zout.writestr(item, zin.read(item.filename))

os.replace(new_zip, OUT)
if os.path.exists(TMP): 
    os.remove(TMP)
print(f"✅ Step3: 注入完成 -> {OUT}")

# ── 5. 验证顺序 ──
with zipfile.ZipFile(OUT) as z:
    final = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

pos_dv  = final.find('<dataValidations')
pos_pm  = final.find('<pageMargins')
pos_ext = final.find('<extLst>')
pos_end = final.find('</worksheet>')
print(f"✅ 顺序验证: dataValidations({pos_dv}) < pageMargins({pos_pm}) < extLst({pos_ext}) < /worksheet({pos_end})")
print(f"   Z列币种={'✅' if '美元,欧元' in final else '❌'}, AA列是否含电={'✅' if '是,否' in final else '❌'}")
