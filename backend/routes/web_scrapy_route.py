"""
web_scrapy_route.py
Flask 路由 — 速卖通订单抓取（完整移植自 aliexpress_scraper.py)

前端 POST /api/web-scrapy/scrape  { "url": "<订单列表URL>", "max_pages": 5 }
返回 xlsx 文件下载

使用方式：
  - 直接调用接口即可，后端会自动检测 Chrome 是否启动
  - 如果 Chrome 未启动，会自动拉起（复用 chrome-selenium profile，保留登录态）
  - 如果是第一次使用，需要先手动在弹出的 Chrome 中登录速卖通，再重试
"""

import os
import time
import socket
import subprocess
from datetime import datetime

from flask import Blueprint, request, jsonify, send_file

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
load_dotenv()

web_scrapy_bp = Blueprint("web_scrapy", __name__)

# ── 常量 ──────────────────────────────────────────────
CHANNEL_ID       = "1579196"
BASE_URL         = "https://csp.aliexpress.com"
PAGE_DELAY       = 3    # 翻页等待秒数
CHROME_PATH      = os.getenv("CHROME_PATH", r"C:\Program Files\Google\Chrome\Application\chrome.exe")
CHROME_USER_DATA = os.getenv("CHROME_USER_DATA", r"C:\Users\ranxi\chrome-selenium")
DEBUG_PORT       = 9222


# ─────────────────────────────────────────────────────
# Driver — 自动检测并启动 Chrome
# ─────────────────────────────────────────────────────

def is_chrome_reachable():
    """检查 Chrome 调试端口是否可连接"""
    try:
        s = socket.create_connection(("127.0.0.1", DEBUG_PORT), timeout=2)
        s.close()
        return True
    except OSError:
        return False


def setup_driver():
    """
    连接已有 Chrome（复用登录态）。
    如果 Chrome 未启动，自动拉起，然后等待端口就绪。
    """
    if not is_chrome_reachable():
        print("🚀 Chrome 未启动，正在自动启动...")
        subprocess.Popen([
            CHROME_PATH,
            f"--remote-debugging-port={DEBUG_PORT}",
            f"--user-data-dir={CHROME_USER_DATA}",
        ])
        # 等待端口就绪，最多 15 秒
        for _ in range(15):
            time.sleep(1)
            if is_chrome_reachable():
                print("✅ Chrome 已就绪")
                break
        else:
            raise RuntimeError(
                "Chrome 启动超时。请手动打开 Chrome 并登录速卖通后重试。"
            )

    options = Options()
    options.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUG_PORT}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(options=options)
    return driver


# ─────────────────────────────────────────────────────
# 列表页解析（第一阶段）+ 详情页抓取（第二阶段）
# ─────────────────────────────────────────────────────

def parse_orders_from_page(driver):
    """
    第一阶段 — 只解析列表数据，不进详情页
    """
    all_orders = []

    # 等待表格加载
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.next-table-row"))
        )
    except Exception:
        print("  [警告] 订单表格未加载")
        return all_orders

    tables = driver.find_elements(By.CSS_SELECTOR, "table.next-table-row")
    print(f"  找到 {len(tables)} 个订单")

    for table in tables:
        order = {}
        try:
            # 订单号 + 构造详情链接
            try:
                order_id_el = table.find_element(
                    By.CSS_SELECTOR,
                    "span.header--valueHighLight--wCk3sLF"
                )
                order_id = order_id_el.text.strip()
                order['order_id'] = order_id
                order['order_link'] = (
                    f"{BASE_URL}/m_apps/order-manage/"
                    f"orderDetail?orderId={order_id}&channelId={CHANNEL_ID}"
                )
            except Exception as e:
                print(f"  [订单号解析失败] {e}")
                order['order_id']   = ""
                order['order_link'] = ""

            # 下单时间
            try:
                time_els = table.find_elements(
                    By.CSS_SELECTOR, "span.header--value--E2HYUZn"
                )
                order['date'] = time_els[0].text.strip() if time_els else ""
            except Exception:
                order['date'] = ""

            # 买家
            try:
                buyer_el = table.find_element(
                    By.CSS_SELECTOR, "a.buyerInfo--inline--U3y4fIR"
                )
                order['buyer'] = buyer_el.text.strip()
            except Exception:
                order['buyer'] = ""

            # 商品名称
            try:
                product_el = table.find_element(
                    By.CSS_SELECTOR, "span.productInfo--itemTitle--QshSnPH"
                )
                order['product'] = product_el.text.strip()[:80]
            except Exception:
                order['product'] = ""

            # 规格 / SKU
            try:
                sku_els = table.find_elements(
                    By.CSS_SELECTOR, "span.productInfo--skuCodeValue--FJA_1Ru"
                )
                order['specs'] = sku_els[0].text.strip() if len(sku_els) > 0 else ""
                order['sku']   = sku_els[1].text.strip() if len(sku_els) > 1 else ""
            except Exception:
                order['specs'] = ""
                order['sku']   = ""

            # 单价
            try:
                price_el = table.find_element(
                    By.CSS_SELECTOR, "span.productInfo--unitFee--mVPKC9G"
                )
                order['price'] = price_el.text.strip()
            except Exception:
                order['price'] = ""

            # 数量
            try:
                qty_el = table.find_element(
                    By.CSS_SELECTOR, "td[data-next-table-col='3'] div"
                )
                order['qty'] = qty_el.text.strip()
            except Exception:
                order['qty'] = ""

            # 总金额
            try:
                amount_el = table.find_element(
                    By.CSS_SELECTOR, "div.amount--amount--YdsJokJ"
                )
                order['amount'] = amount_el.text.strip()
            except Exception:
                order['amount'] = ""

            # 订单状态
            try:
                status_el = table.find_element(
                    By.CSS_SELECTOR, "div.chc-state-label__stateText"
                )
                order['status']    = status_el.text.strip()
                order['status_en'] = translate_status(order['status'])
            except Exception:
                order['status']    = ""
                order['status_en'] = ""

            # AE/IOSS
            try:
                ioss_els   = table.find_elements(By.CSS_SELECTOR, "span.chc-color-tag")
                ioss_texts = [el.text for el in ioss_els]
                order['ae_ioss'] = "yes" if "AE/IOSS" in ioss_texts else "no"
            except Exception:
                order['ae_ioss'] = "no"

            # 半托管
            try:
                tag_els    = table.find_elements(By.CSS_SELECTOR, "span.chc-color-tag")
                tag_texts  = [el.text for el in tag_els]
                order['semi_managed'] = "yes" if any("半托管" in t for t in tag_texts) else "no"
            except Exception:
                order['semi_managed'] = "no"

            # 操作按钮
            try:
                btns = table.find_elements(
                    By.CSS_SELECTOR, "button.next-btn span.next-btn-helper"
                )
                btn_texts    = [b.text.strip() for b in btns if b.text.strip()]
                order['action'] = ", ".join(btn_texts)
            except Exception:
                order['action'] = ""

            # 详情字段先占位
            order['recipient']   = ""
            order['address']     = ""
            order['postal_code'] = ""
            order['email']       = ""
            order['phone']       = ""
            order['tax_number']  = ""

            all_orders.append(order)

        except Exception as e:
            print(f"  [跳过] 解析订单出错: {e}")
            continue

    return all_orders



def extract_order_detail(driver, order_link):
    """进入详情页，点击完整收货地址眼睛，提取收货信息"""
    result = {
        'recipient':   '',
        'address':     '',
        'postal_code': '',
        'email':       '',
        'phone':       '',
        'tax_number':  ''
    }

    try:
        driver.get(order_link)

        # 等待地址区域加载
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div[class*='orderInfo--addressItem']")
                )
            )
        except Exception:
            print("  [详情页] 地址区域加载超时，跳过此订单")
            return result  # 返回空的 result 字典即可

        # ── 收集地址区域所有元素，返回给前端用于定位按钮 ──
        debug_elements = []
        try:
            container = driver.find_element(
                By.XPATH, "//*[contains(@class,'orderInfo--address')]"
            )
            for el in container.find_elements(By.XPATH, ".//*")[:50]:
                tag = el.tag_name
                cls = el.get_attribute("class") or ""
                txt = (el.text or "").strip()[:40]
                onclick = el.get_attribute("onclick") or ""
                if cls or txt:
                    debug_elements.append({
                        "tag": tag, "class": cls, "text": txt, "onclick": onclick
                    })
        except Exception as e:
            debug_elements.append({"error": str(e)})

        # ── 尝试多种方式点击展开按钮 ──
        # clicked_by = None
        # 页面有两个眼睛图标：左边是买家名字旁，右边是完整收货地址
        # 右边收货地址的眼睛 data-spm-anchor-id 包含 "i3"
        # clicked_by = None
        try:
            eye_els = driver.find_elements(By.CSS_SELECTOR, "i[class*='orderEye--eye']")
            target = None
            for el in eye_els:
                spm = el.get_attribute("data-spm-anchor-id") or ""
                if ".i3." in spm:
                    target = el
                    break
            # 如果没找到 i3，fallback 取最后一个（通常右边）
            if target is None and eye_els:
                target = eye_els[-1]
            if target:
                driver.execute_script("arguments[0].click();", target)
                # clicked_by = "i[class*='orderEye--eye'][data-spm=i3]"
                print("     OK 点击收货地址眼睛")
                time.sleep(1)
            else:
                print("    WARN 未找到眼睛按钮")
        except Exception as e:
            print(f"    WARN 点击眼睛失败: {e}")

        # ── 等收件人脱敏 ──
        # unmasked = False
        def recipient_unmasked(d):
            items = d.find_elements(By.CSS_SELECTOR, "div[class*='orderInfo--addressItem']")
            for item in items:
                try:
                    label = item.find_element(By.CSS_SELECTOR, "span[class*='addressLabel']").text.strip()
                    if "收件人名称" in label:
                        value = item.find_element(By.CSS_SELECTOR, "span[class*='addressValue']").text.strip()
                        return "*" not in value and value != ""
                except Exception:
                    pass
            return False

        try:
            WebDriverWait(driver, 5).until(recipient_unmasked)
            print("    OK 收件人已脱敏")
        except Exception:
            # 等待超时就兜底等 2 秒再读，避免空结果
            print("    WARN 等待收件人脱敏超时，延迟2秒后继续")
            time.sleep(2)


        # ── 读取地址字段 ──
        result = {
            'recipient': '', 'address': '', 'postal_code': '',
            'email': '', 'phone': '', 'tax_number': ''
        }
        address_items = driver.find_elements(By.CSS_SELECTOR, "div[class*='orderInfo--addressItem']")
        for item in address_items:
            try:
                label = item.find_element(By.CSS_SELECTOR, "span[class*='addressLabel']").text.strip()
                value = item.find_element(By.CSS_SELECTOR, "span[class*='addressValue']").text.strip()
                if "收件人名称" in label:
                    result['recipient'] = value
                elif "详细地址" in label:
                    result['address'] = value
                elif "邮编" in label:
                    result['postal_code'] = value
                elif "联系邮件" in label:
                    result['email'] = value
                elif "联系电话" in label:
                    result['phone'] = value
                elif "Tax" in label:
                    result['tax_number'] = value
            except Exception:
                continue
        print(f"recipient: {result['recipient']}")
        print(f"address: {result['address']}")
        print(f"postal_code: {result['postal_code']}")
        print(f"email: {result['email']}")
        print(f"phone: {result['phone']}")
        print(f"tax_number: {result['tax_number']}")

    except Exception as e:
        print(f"  [详情页错误] {e}")

    return result

# ─────────────────────────────────────────────────────
# 翻页
# ─────────────────────────────────────────────────────


def get_total_pages(driver):
    """从分页显示元素读取总页数，例如 '1/30' 返回 30"""
    try:
        display = driver.find_element(By.CSS_SELECTOR, "span.next-pagination-display")
        text = display.text.strip()  # 例如 "1/30"
        total = int(text.split("/")[-1])
        print(f"  [分页] 当前: {text}，共 {total} 页")
        return total
    except Exception as e:
        print(f"  [分页] 读取总页数失败: {e}")
        return None


def go_next_page(driver, current_page):
    """点击下一页，返回 True 表示成功翻页，False 表示已是最后一页"""
    try:
        total = get_total_pages(driver)
        if total is not None and current_page >= total:
            print(f"  [翻页] 已是最后一页 ({current_page}/{total})")
            return False

        next_btn = driver.find_element(
            By.CSS_SELECTOR, "button.next-pagination-item.next-next"
        )
        disabled = next_btn.get_attribute("disabled")
        aria_label = next_btn.get_attribute("aria-label") or ""
        print(f"  [翻页] 找到下一页按钮: aria-label='{aria_label}', disabled='{disabled}'")
        if disabled is not None:
            print("  [翻页] 按钮已禁用，已是最后一页")
            return False
        driver.execute_script("arguments[0].click();", next_btn)
        print("  [翻页] 点击下一页成功")
        time.sleep(PAGE_DELAY)
        return True
    except Exception as e:
        print(f"  [翻页] 未找到下一页按钮，停止: {e}")
        return False


# ─────────────────────────────────────────────────────
# 状态翻译
# ─────────────────────────────────────────────────────

def translate_status(status):
    mapping = {
        '等待发货':     'Awaiting shipment',
        '等待买家收货': 'Awaiting buyer receipt',
        '交易成功':     'Transaction complete',
        '已关闭':       'Closed',
        '等待付款':     'Awaiting payment',
        '等待买家付款': 'Awaiting payment',
        '等待仓库发货': 'Awaiting warehouse shipment',
    }
    return mapping.get(status, status)


# ─────────────────────────────────────────────────────
# 保存 Excel
# ─────────────────────────────────────────────────────

def save_orders_to_xlsx(all_orders):
    base_dir     = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(base_dir, "..", "downloads")
    os.makedirs(download_dir, exist_ok=True)

    filename = "order_list.xlsx"
    filepath = os.path.join(download_dir, filename)

    headers = [
        "Order ID", "Order Link", "Date", "Buyer",
        "Product", "Specs", "SKU", "Price", "Qty", "Amount",
        "Status (中文)", "Status (EN)", "AE/IOSS", "Semi-Managed", "Action",
        "Recipient", "Address", "Postal Code", "Email", "Phone", "Tax Number"
    ]

    # ── 读取已有文件中的订单（用于去重）──────────────────
    existing_orders = {}
    if os.path.exists(filepath):
        try:
            from openpyxl import load_workbook
            wb_old = load_workbook(filepath)
            ws_old = wb_old.active
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                order_id = str(row[0]).lstrip("'").strip() if row[0] else ""
                if order_id:
                    row = list(row)
                    # 统一把日期列转为 datetime 对象
                    if row[2] and not isinstance(row[2], datetime):
                        try:
                            row[2] = datetime.strptime(str(row[2]), "%Y-%m-%d %H:%M:%S")
                        except Exception:
                            pass
                    existing_orders[order_id] = row
            print(f"  读取已有文件，共 {len(existing_orders)} 条历史订单")
        except Exception as e:
            print(f"  读取已有文件失败，将覆盖: {e}")

    # ── 新数据合并，order_id 相同则用新数据覆盖 ──────────
    for order in all_orders:
        order_id = order.get('order_id', '').strip()
        if order_id:
            # 将日期字符串转为 datetime 对象，方便排序和 Excel 格式化
            raw_date = order.get('date', '')
            try:
                parsed_date = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S")
            except Exception:
                parsed_date = None

            existing_orders[order_id] = [
                "'" + order.get('order_id', ''),
                order.get('order_link', ''),
                parsed_date,                        # datetime 对象，Excel 会自动格式化
                order.get('buyer', ''),
                order.get('product', ''),
                order.get('specs', ''),
                order.get('sku', ''),
                order.get('price', ''),
                order.get('qty', ''),
                order.get('amount', ''),
                order.get('status', ''),
                order.get('status_en', ''),
                order.get('ae_ioss', ''),
                order.get('semi_managed', ''),
                order.get('action', ''),
                order.get('recipient', ''),
                order.get('address', ''),
                order.get('postal_code', ''),
                order.get('email', ''),
                order.get('phone', ''),
                order.get('tax_number', ''),
            ]

    # ── 按日期降序排列（最新的在最上面）─────────────────
    def sort_key(row):
        date_val = row[2]
        if isinstance(date_val, datetime):
            return date_val
        return datetime.min

    sorted_rows = sorted(existing_orders.values(), key=sort_key, reverse=True)

    # ── 写入文件 ──────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, values in enumerate(sorted_rows, 2):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            # 日期列设置格式
            if col == 3 and isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

        # Order Link 设为超链接
        link_val = values[1] if len(values) > 1 else ""
        if link_val:
            ws.cell(row=row_idx, column=2).hyperlink = link_val
            ws.cell(row=row_idx, column=2).font = Font(color="0563C1", underline="single")

    col_widths = [20, 15, 18, 12, 40, 25, 15, 12, 6, 12, 16, 20, 8, 14, 20, 20, 35, 12, 25, 15, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(filepath)
    print(f"OK 已保存 {len(existing_orders)} 条订单（含历史）-> {filepath}")
    return filepath, filename


# ─────────────────────────────────────────────────────
# 主抓取流程
# ─────────────────────────────────────────────────────

def crawl_orders(order_list_url, max_pages=None):
    # 连接已有 Chrome，绝对不能调 driver.quit()，否则会关掉用户的浏览器
    driver = setup_driver()

    print(f"Opening order list: {order_list_url}")
    driver.get(order_list_url)

    # 等页面真正加载完（等到有订单表格出现，最多 30 秒）
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.next-table-row"))
        )
        print("OK 订单列表页已加载")
    except Exception:
        print("WARN 等待订单列表超时，尝试继续...")

    # 等待用户在浏览器里点击查询按钮
    print("⏳ 请在浏览器里点击【查询】按钮，程序将自动继续...")
    try:
        # 找到查询按钮，注入 JS 监听点击事件，点击后设置一个标记
        query_btn = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.XPATH, "//button[.//span[text()='查询']]")
            )
        )
        driver.execute_script("""
            arguments[0].addEventListener('click', function() {
                window.__query_clicked = true;
            });
        """, query_btn)
        print("  监听查询按钮中，等待用户点击...")

        # 等用户点击（最多 5 分钟）
        WebDriverWait(driver, 300).until(
            lambda d: d.execute_script("return window.__query_clicked === true;")
        )
        print("OK 检测到用户点击查询，等待列表刷新...")
        time.sleep(2)

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.next-table-row"))
        )
        print("OK 订单列表已加载，开始抓取...")
        time.sleep(1)
    except Exception as e:
        print(f"WARN 等待查询超时，尝试继续: {e}")

    all_orders = []
    page = 1

    # ── 第一阶段：翻完所有页，只收集列表数据 ──────────────
    while True:
        print(f"\n第 {page} 页（收集列表）...")
        orders = parse_orders_from_page(driver)
        all_orders.extend(orders)
        print(f"  累计 {len(all_orders)} 条")

        if max_pages and page >= max_pages:
            print(f"  已达设定最大页数 {max_pages}，停止翻页")
            break

        if not go_next_page(driver, page):
            print("  已到最后一页")
            break

        page += 1

        # 每 5 页额外冷却
        if page % 5 == 0:
            print("  冷却 10 秒...")
            time.sleep(10)

    # ── 第二阶段：统一抓所有详情页（不再需要回列表页）──────
    print(f"\n开始抓取 {len(all_orders)} 条订单的详情页...")
    for order in all_orders:
        if order.get("order_link"):
            print(f"  -> 抓详情 {order['order_id']}")
            detail_data = extract_order_detail(driver, order["order_link"])
            order.update(detail_data)

    # 不调用 driver.quit()，Chrome 是用户自己的，不能关
    return all_orders

# ─────────────────────────────────────────────────────
# Flask 路由
# ─────────────────────────────────────────────────────

@web_scrapy_bp.route("/api/web-scrapy/scrape", methods=["POST", "OPTIONS"])
def scrape_web_page():
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200

    data      = request.get_json()
    url       = data.get("url")
    max_pages = data.get("max_pages", None)  # 可选，不传则抓全部

    if not url:
        return jsonify({"error": "URL is required"}), 400

    print(f"Start scraping: {url}, max_pages: {max_pages or 'ALL'}")

    try:
        all_orders = crawl_orders(url, max_pages=max_pages)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503

    if not all_orders:
        return jsonify({"error": "No orders scraped"}), 400

    xlsx_path, xlsx_name = save_orders_to_xlsx(all_orders)
    print(f"File saved: {xlsx_path}")

    return send_file(
        xlsx_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=xlsx_name
    )




@web_scrapy_bp.route("/api/web-scrapy/scrape-detail", methods=["POST", "OPTIONS"])
def scrape_order_detail():
    """
    单条订单详情调试接口
    POST { "url": "https://csp.aliexpress.com/m_apps/order-manage/orderDetail?orderId=xxx&channelId=xxx" }
    返回 JSON，包含收货信息 + 调试用的页面元素列表
    """
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200

    data = request.get_json()
    url  = data.get("url")

    if not url:
        return jsonify({"error": "URL is required"}), 400

    print(f"[detail] 开始抓单条详情: {url}")

    try:
        driver = setup_driver()
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503

    try:
        driver.get(url)

        # 等待地址区域加载
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div[class*='orderInfo--addressItem']")
                )
            )
        except Exception:
            return jsonify({"error": "页面加载超时，地址区域未出现，请确认已登录速卖通"}), 504

        # ── 收集地址区域所有元素，返回给前端用于定位按钮 ──
        debug_elements = []
        try:
            container = driver.find_element(
                By.XPATH, "//*[contains(@class,'orderInfo--address')]"
            )
            for el in container.find_elements(By.XPATH, ".//*")[:50]:
                tag = el.tag_name
                cls = el.get_attribute("class") or ""
                txt = (el.text or "").strip()[:40]
                onclick = el.get_attribute("onclick") or ""
                if cls or txt:
                    debug_elements.append({
                        "tag": tag, "class": cls, "text": txt, "onclick": onclick
                    })
        except Exception as e:
            debug_elements.append({"error": str(e)})

        # ── 点击收货地址眼睛（用 data-spm i3 定位，避免误点买家名字旁的眼睛）──
        clicked_by = None
        try:
            eye_els = driver.find_elements(By.CSS_SELECTOR, "i[class*='orderEye--eye']")
            target = None
            for el in eye_els:
                spm = el.get_attribute("data-spm-anchor-id") or ""
                if ".i3." in spm:
                    target = el
                    break
            if target is None and eye_els:
                target = eye_els[-1]
            if target:
                driver.execute_script("arguments[0].click();", target)
                clicked_by = "orderEye--eye[data-spm=i3]"
                print("    OK 点击收货地址眼睛")
                time.sleep(1)
            else:
                print("    WARN 未找到眼睛按钮")
        except Exception as e:
            print(f"    WARN 点击眼睛失败: {e}")

        # ── 等收件人脱敏 ──
        unmasked = False
        def recipient_unmasked(d):
            items = d.find_elements(By.CSS_SELECTOR, "div[class*='orderInfo--addressItem']")
            for item in items:
                try:
                    label = item.find_element(By.CSS_SELECTOR, "span[class*='addressLabel']").text.strip()
                    if "收件人名称" in label:
                        value = item.find_element(By.CSS_SELECTOR, "span[class*='addressValue']").text.strip()
                        return "*" not in value and value != ""
                except Exception:
                    pass
            return False

        try:
            WebDriverWait(driver, 30).until(recipient_unmasked)
            print("    OK 收件人已脱敏")
        except Exception:
            # 等待超时就兜底等 2 秒再读，避免空结果
            print("    WARN 等待收件人脱敏超时，延迟2秒后继续")
            time.sleep(3)


        # ── 读取地址字段 ──
        result = {
            'recipient': '', 'address': '', 'postal_code': '',
            'email': '', 'phone': '', 'tax_number': ''
        }
        address_items = driver.find_elements(By.CSS_SELECTOR, "div[class*='orderInfo--addressItem']")
        for item in address_items:
            try:
                label = item.find_element(By.CSS_SELECTOR, "span[class*='addressLabel']").text.strip()
                value = item.find_element(By.CSS_SELECTOR, "span[class*='addressValue']").text.strip()
                if "收件人名称" in label:
                    result['recipient'] = value
                elif "详细地址" in label:
                    result['address'] = value
                elif "邮编" in label:
                    result['postal_code'] = value
                elif "联系邮件" in label:
                    result['email'] = value
                elif "联系电话" in label:
                    result['phone'] = value
                elif "Tax" in label:
                    result['tax_number'] = value
            except Exception:
                continue
        print(f"recipient: {result['recipient']}")
        print(f"address: {result['address']}")
        print(f"postal_code: {result['postal_code']}")
        print(f"email: {result['email']}")
        print(f"phone: {result['phone']}")
        print(f"tax_number: {result['tax_number']}")
        
        return jsonify({
            "data":           result,
            "unmasked":       unmasked,
            "clicked_by":     clicked_by,
            "debug_elements": debug_elements,   # 用这个找正确的按钮 selector
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500
