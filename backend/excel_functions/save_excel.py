from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict
import copy
import os
import re
import zipfile
# ─────────────────────────────────────────────────────
# 保存 Excel
# ─────────────────────────────────────────────────────
def write_excel(filename: str,filepath:str, excel_headers: List[str], all_rows, col_widths:List[int]):
    
    wb = Workbook()
    ws = wb.active
    ws.title = filename

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(excel_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(all_rows, 2):

        for col_idx, value in enumerate(row, 1):

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

        link_val = row[-1]

        if link_val:
            ws.cell(row=row_idx, column=len(row)).hyperlink = link_val
            ws.cell(row=row_idx, column=len(row)).font = Font(color="0563C1", underline="single")

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(filepath)

    print(f"OK 已保存 {len(all_rows)} 条数据 -> {filepath}")

    wb.close()
    
def save_orders_to_xlsx(data:List[Dict[str, str]], filename:str, data_keys:List[str],excel_headers:List[str], col_widths:List[int], mode:str='a'):

    base_dir = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(base_dir, "..", "..", "..", "downloads")
    os.makedirs(download_dir, exist_ok=True)

    
    filepath = os.path.join(download_dir, filename)

    old_rows = []

    # ─————————————─ 如果文件存在，读取旧数据 ─────────────────────────────
    if mode=='a' and os.path.exists(filepath):

        wb_old = load_workbook(filepath)
        ws_old = wb_old.active

        for row in ws_old.iter_rows(min_row=2, values_only=True):
            old_rows.append(list(row))

        wb_old.close()

        print(f"读取旧Excel {len(old_rows)} 条")

    # ──—————————— 处理新订单 ─────────────────────────────
    
    new_rows: List[List[str]]= []

    for order in data:
        # print(f"[save_orders_to_xlsx]Order row is {order}")
        raw_date = order.get("date", "")

        
        new_row=[]
        
        for key in data_keys:
            if key=='date':
                parsed_date = None
                try:
                    parsed_date = datetime.strptime(raw_date, "%m/%d/%Y %H:%M")
                except Exception:
                    pass
                new_row.append(parsed_date)
                continue
            new_row.append(order.get(key, ''))
        new_rows.append(new_row)
      
    # ─————————─ ✅ 模式控制 ─────────────────────────────
    all_rows = None
    if mode=='a':
        all_rows = new_rows + old_rows
    elif mode == "w":
        all_rows = new_rows
    else:
        raise ValueError("mode must be 'a' or 'w'")    
    
    # ──———————— 写 Excel ─────────────────────────────
    write_excel(filename=filename, filepath=filepath, excel_headers=excel_headers, all_rows=all_rows, col_widths=col_widths)

    return filepath, filename

# ─────────────────────────────────────────────────────
# 燕文模板写入（保留模板格式 + 下拉）
# ─────────────────────────────────────────────────────
def _safe_copy_style(src, tgt):
    for attr in ("font", "fill", "border", "alignment"):
        try:
            v = getattr(src, attr)
            if v:
                setattr(tgt, attr, copy.copy(v))
        except Exception:
            pass
    try:
        if src.number_format:
            tgt.number_format = src.number_format
    except Exception:
        pass
 
def save_yanwen_to_xlsx(
    data: List[Dict[str, str]],
    data_keys: List[str],
    filename: str,
    column_widths: List[int],
):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(base_dir, "..", "..", "..", "downloads")
    os.makedirs(download_dir, exist_ok=True)
 
    filepath = os.path.join(download_dir, filename)
    tmp_path = filepath + ".tmp.xlsx"
 
    # 模板路径（与 save_excel.py 同目录）
    template_path = os.path.join(base_dir, "Yanwen小包专线模板_V251224.xlsx")
 
    total_rows = len(data)
    template_row = 2
 
    # ── Step 1: openpyxl 写数据 + 样式 ──────────────────
    wb = load_workbook(template_path)
    ws = wb.active
    ws.data_validations.dataValidation.clear()  # 先清掉，后面用 XML 注入
 
    # 复制模板行样式到每个数据行
    for i in range(template_row, template_row + total_rows):
        for col in range(1, ws.max_column + 1):
            _safe_copy_style(
                ws.cell(row=template_row, column=col),
                ws.cell(row=i, column=col)
            )
 
    # 填入数据
    for row_idx, order in enumerate(data, start=template_row):
        for col_idx, key in enumerate(data_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=order.get(key, ""))
 
    # 设置列宽
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
 
    wb.save(tmp_path)
    print("✅ Step1: 数据写入完成")
 
    # ── Step 2: 从原模板提取验证块 ──────────────────────
    with zipfile.ZipFile(template_path) as z:
        orig_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
 
    std_dv  = re.search(r'<dataValidations\b.*?</dataValidations>', orig_xml, re.DOTALL)
    ext_lst = re.search(r'<extLst>.*?</extLst>', orig_xml, re.DOTALL)
 
    std_dv_xml  = std_dv.group(0)  if std_dv  else ""
    ext_lst_xml = ext_lst.group(0) if ext_lst else ""
    print(f"✅ Step2: 标准验证={'找到' if std_dv_xml else '无'}, extLst={'找到' if ext_lst_xml else '无'}")
 
    # ── Step 3: 按正确顺序注入 XML ──────────────────────
    with zipfile.ZipFile(tmp_path) as z:
        tmp_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
 
    # dataValidations 插到 <hyperlinks / <pageMargins / <pageSetup / </worksheet> 之前
    anchor_tags = ['<hyperlinks', '<pageMargins', '<pageSetup', '</worksheet>']
    anchor_pos = len(tmp_xml)
    anchor_tag_used = '</worksheet>'
    for tag in anchor_tags:
        idx = tmp_xml.find(tag)
        if idx != -1 and idx < anchor_pos:
            anchor_pos = idx
            anchor_tag_used = tag
 
    print(f"  dataValidations 插入锚点: '{anchor_tag_used}'")
 
    if std_dv_xml:
        tmp_xml = tmp_xml[:anchor_pos] + std_dv_xml + tmp_xml[anchor_pos:]
 
    # extLst 插到 </worksheet> 前
    if ext_lst_xml:
        tmp_xml = tmp_xml.replace("</worksheet>", ext_lst_xml + "</worksheet>")
 
    # ── Step 4: 重新打包 ────────────────────────────────
    new_zip = filepath + ".new"
    with zipfile.ZipFile(tmp_path, 'r') as zin, \
         zipfile.ZipFile(new_zip, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == "xl/worksheets/sheet1.xml":
                zout.writestr(item, tmp_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
 
    os.replace(new_zip, filepath)
    if os.path.exists(tmp_path):
        os.remove(tmp_path)
 
    print(f"OK 燕文模板已保存 {total_rows} 条 -> {filepath}")
 
    return filepath, filename