"""
get_dropdown_values.py

完整提取 Yanwen小包专线模板_V251224.xlsx 的所有下拉列配置。

模板中共有两种下拉：
  1. 标准 dataValidation（openpyxl 可直接读）
       Z  列：币种类型   → 直接内嵌字符串列表
       AA 列：是否含电   → 直接内嵌字符串列表

  2. x14 扩展验证 ExtLst（openpyxl 不支持，需读参考 sheet）
       C  列：交货仓     → 交货仓!$B$2:$B$1000
       D  列：产品名称   → 燕文产品名称!$B$2:$B$1000
       J  列：收件人国家 → 国家中英文&二字代码!$B$2:$B$1000

返回统一格式：
  {
    "C":  ["北京燕文", "青岛燕文", ...],   # 30 条
    "D":  ["大陆DHL", "燕文英国YODEL...", ...],  # 178 条
    "J":  ["阿鲁巴", "格陵兰岛", ...],    # 248 条
    "Z":  ["美元", "欧元", ...],           # 19 条
    "AA": ["是", "否"],                    # 2 条
  }
"""


from openpyxl import load_workbook
# from openpyxl.worksheet.datavalidation import DataValidation
import warnings
import zipfile
import re
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ─────────────────────────────────────────────────────────────
# 内部工具
# ─────────────────────────────────────────────────────────────

def _read_sheet_col(wb, sheet_name: str, col: int = 2) -> list:
    """从参考 sheet 读取某列的非空值列表"""
    try:
        ws = wb[sheet_name]
    except KeyError:
        print(f"⚠️  Sheet '{sheet_name}' 不存在，跳过")
        return []
    return [
        ws.cell(r, col).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, col).value is not None
    ]


def _parse_inline_list(formula1: str) -> list:
    """把 '"美元,欧元,英镑"' 解析为 ['美元', '欧元', '英镑']"""
    s = formula1.strip().strip('"')
    return [v.strip() for v in s.split(",") if v.strip()]


def _parse_x14_validations(xlsx_path: str) -> dict:
    """
    直接读 sheet1.xml 里的 x14:dataValidation，
    返回 {sqref首列字母: sheet引用} 的映射，例如：
      {"C": "交货仓!$B$2:$B$1000", "D": "燕文产品名称!$B$2:$B$1000", ...}
    """
    result = {}
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    except Exception as e:
        print(f"⚠️  读取 sheet1.xml 失败: {e}")
        return result

    blocks = re.findall(
        r'<x14:dataValidation.*?</x14:dataValidation>', xml, re.DOTALL
    )
    for block in blocks:
        sqref_m  = re.search(r'<xm:sqref>(.*?)</xm:sqref>', block)
        formula_m = re.search(r'<xm:f>(.*?)</xm:f>', block)
        if not sqref_m or not formula_m:
            continue

        # 取 sqref 第一段的列字母，例如 "C1:C6 C8:C1048576" → "C"
        sqref   = sqref_m.group(1).strip()
        col_letter = re.match(r'([A-Z]+)', sqref)
        if not col_letter:
            continue

        formula = formula_m.group(1).strip()
        # 还原 XML 转义
        formula = formula.replace("&amp;", "&")

        result[col_letter.group(1)] = formula

    return result


# ─────────────────────────────────────────────────────────────
# 主函数
# ─────────────────────────────────────────────────────────────

def get_all_dropdown_values(xlsx_path: str) -> dict:
    """
    提取模板中所有下拉列的完整选项。

    Returns:
        dict[col_letter -> list[str]]
        例：{"C": [...], "D": [...], "J": [...], "Z": [...], "AA": [...]}
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active
    dropdown_map = {}

    # ── 1. 标准 dataValidation（type="list"）──────────────────
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not dv.formula1:
            continue

        formula = dv.formula1.strip()

        # 内嵌字符串，例如 "美元,欧元,..."
        if formula.startswith('"') and "!" not in formula:
            values = _parse_inline_list(formula)
            if not values:
                continue

            # 从 sqref 提取所有涉及的列字母
            for segment in str(dv.sqref).split():
                col_letter = re.match(r'([A-Z]+)', segment)
                if col_letter:
                    col = col_letter.group(1)
                    if col not in dropdown_map:
                        dropdown_map[col] = values
                        print(f"✅ 标准验证  {col:3s} | {len(values):4d} 条 | {values[:3]}...")

        # Sheet 引用（标准格式，openpyxl 能读到的）
        elif "!" in formula:
            formula_clean = formula.strip("'\"").replace("&amp;", "&")
            # 解析 SheetName!$B$2:$B$N
            m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?[A-Z]+\$?(\d+)", formula_clean)
            if m:
                sheet_name, col_alpha, row_start, row_end = m.groups()
                col_idx = ord(col_alpha) - ord('A') + 1
                values = _read_sheet_col(wb, sheet_name, col_idx)
                if values:
                    for segment in str(dv.sqref).split():
                        col_letter = re.match(r'([A-Z]+)', segment)
                        if col_letter:
                            col = col_letter.group(1)
                            if col not in dropdown_map:
                                dropdown_map[col] = values
                                print(f"✅ 标准验证  {col:3s} | {len(values):4d} 条 | sheet='{sheet_name}'")

    # ── 2. x14 扩展验证 ExtLst（openpyxl 不支持，直接读 XML）────
    x14_map = _parse_x14_validations(xlsx_path)

    for col_letter, formula in x14_map.items():
        if col_letter in dropdown_map:
            continue  # 已经被标准验证覆盖，跳过

        # 解析 sheet 引用，例如 "交货仓!$B$2:$B$1000"
        m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?\d+", formula)
        if not m:
            print(f"⚠️  x14 {col_letter:3s} 无法解析公式: {formula}")
            continue

        sheet_name, col_alpha = m.groups()
        col_idx = ord(col_alpha) - ord('A') + 1
        values = _read_sheet_col(wb, sheet_name, col_idx)

        if values:
            dropdown_map[col_letter] = values
            print(f"✅ x14 验证  {col_letter:3s} | {len(values):4d} 条 | sheet='{sheet_name}'")
        else:
            print(f"⚠️  x14 {col_letter:3s} | sheet='{sheet_name}' 无数据")

    return dropdown_map


def get_dropdown_values(xlsx_path: str, col_letter: str, keyword: str = None) -> list:
    """
    兼容旧接口：获取单列的下拉值。
    keyword 参数保留但已不需要（直接从 sheet 读准确值）。
    """
    all_map = get_all_dropdown_values(xlsx_path)
    values = all_map.get(col_letter.upper(), [])
    if keyword:
        values = [v for v in values if keyword in str(v)]
    return values


# ─────────────────────────────────────────────────────────────
# 测试入口
# ─────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     path = "Yanwen小包专线模板_V251224.xlsx"
#     print(f"分析文件: {path}\n")

#     all_dropdowns = get_all_dropdown_values(path)

#     print(f"\n{'='*50}")
#     print(f"共发现 {len(all_dropdowns)} 个下拉列\n")

#     for col, vals in sorted(all_dropdowns.items()):
#         print(f"  {col:3s} | {len(vals):4d} 条 | 前3: {vals[:3]}")