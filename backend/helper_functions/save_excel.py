from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
# import time


load_dotenv()
# ─────────────────────────────────────────────────────
# 保存 Excel
# ─────────────────────────────────────────────────────


def save_orders_to_xlsx(data, filename, data_keys,excel_headers, col_widths):

    base_dir = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(base_dir, "..", "..", "..", "downloads")
    os.makedirs(download_dir, exist_ok=True)

    
    filepath = os.path.join(download_dir, filename)

    old_rows = []

    # ── 如果文件存在，读取旧数据 ─────────────────────────────
    if os.path.exists(filepath):

        from openpyxl import load_workbook

        wb_old = load_workbook(filepath)
        ws_old = wb_old.active

        for row in ws_old.iter_rows(min_row=2, values_only=True):
            old_rows.append(list(row))

        wb_old.close()

        print(f"读取旧Excel {len(old_rows)} 条")

    # ── 处理新订单 ─────────────────────────────
    new_rows = []

    for order in data:
        # print(f"[save_orders_to_xlsx]Order row is {order}")
        raw_date = order.get("date", "")

        
        new_row=[]
        
        for key in data_keys:
            if key=='date':
                parsed_date = None
                try:
                    parsed_date = datetime.strptime(raw_date, "%m/%d/%Y %H:%M")
                except Exception:
                    pass
                new_row.append(parsed_date)
                continue
            new_row.append(order.get(key, ''))
        new_rows.append(new_row)
      
    # ── 新订单在前，旧订单在后 ─────────────────────────────
    all_rows = new_rows + old_rows

    # ── 写 Excel ─────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(excel_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(all_rows, 2):

        for col_idx, value in enumerate(row, 1):

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

        link_val = row[-1]

        if link_val:
            ws.cell(row=row_idx, column=len(row)).hyperlink = link_val
            ws.cell(row=row_idx, column=len(row)).font = Font(color="0563C1", underline="single")

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(filepath)

    print(f"OK 已保存 {len(all_rows)} 条数据 -> {filepath}")

    wb.close()

    return filepath, filename

