from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from utils.web_scrape_constant_values import mapping, profile_map, DEBUG_PORT
import os
import time
import socket

load_dotenv()


# ─────────────────────────────────────────────────────
# Driver — 自动检测并启动 Chrome
# ─────────────────────────────────────────────────────

def is_chrome_reachable():
    """检查 Chrome 调试端口是否可连接"""
    try:
        s = socket.create_connection(("127.0.0.1", DEBUG_PORT), timeout=2)
        s.close()
        return True
    except OSError:
        return False
# ─────────────────────────────────────────────────────
# 状态翻译
# ─────────────────────────────────────────────────────

def translate_status(status):
    
    return mapping.get(status, status)



def setup_driver(channel_id):

    

    profile_name = profile_map.get(channel_id)

    if not profile_name:
        raise RuntimeError(f"Unknown channel id {channel_id}")

    profile_root = os.path.join(os.getcwd(), "chrome_profiles")
    os.makedirs(profile_root, exist_ok=True)

    profile_dir = os.path.join(profile_root, profile_name)
    os.makedirs(profile_dir, exist_ok=True)

    options = Options()

    options.add_argument(f"--user-data-dir={profile_dir}")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(options=options)

    return driver

def get_driver(channel_id, driver_pool):

    if channel_id not in driver_pool:
        print(f"Initializing driver for channel {channel_id}")
        driver_pool[channel_id] = setup_driver(channel_id)

    return driver_pool[channel_id]
# ─────────────────────────────────────────────────────
# 保存 Excel
# ─────────────────────────────────────────────────────

def save_orders_to_xlsx(all_orders, store):

    base_dir     = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(base_dir, "..", "..", "downloads")
    os.makedirs(download_dir, exist_ok=True)

    filename = "order_list.xlsx"
    filepath = os.path.join(download_dir, filename)

    headers = [
        "Store",  "Order ID",  "Date", "Buyer",
        "Product", "Specs", "SKU", "Price", "Qty", "Amount",
        "Status (中文)", "Status (EN)", "AE/IOSS", "Semi-Managed", "Action",
        "Recipient", "Address", "Postal Code", "Email", "Phone", "Tax Number", "Order Link"
    ]

    existing_orders = {}

    # ── 读取已有文件中的订单（用于去重）──────────────────
    if os.path.exists(filepath):
        try:
            from openpyxl import load_workbook
            wb_old = load_workbook(filepath)
            ws_old = wb_old.active

            for row in ws_old.iter_rows(min_row=2, values_only=True):

                # 新结构：store 在第0列，order_id 在第1列
                order_id = str(row[1]).lstrip("'").strip() if row[1] else ""

                if order_id:
                    row = list(row)

                    if row[2] and not isinstance(row[2], datetime):
                        try:
                            row[2] = datetime.strptime(str(row[2]), "%Y-%m-%d %H:%M:%S")
                        except Exception:
                            pass

                    existing_orders[order_id] = row

            print(f"  读取已有文件，共 {len(existing_orders)} 条历史订单")

        except Exception as e:
            print(f"  读取已有文件失败，将覆盖: {e}")

    # ── 新数据合并 ───────────────────────────────────
    for order in all_orders:

        order_id = order.get('order_id', '').strip()

        if order_id is None:
            continue

        raw_date = order.get('date', '')

        # print(f"[save orders to xlsx]Raw date is {raw_date}")
        # try:
        #     # 如果是纯数字（时间戳）
        #     if raw_date.isdigit():
        #             ts = int(raw_date)

        #             # 16位 → 微秒
        #             if len(raw_date) >= 16:
        #                 ts = ts / 1_000_000
        #             # 13位 → 毫秒
        #             elif len(raw_date) == 13:
        #                 ts = ts / 1000

        #             parsed_date = datetime.fromtimestamp(ts)

        #     else:
        #             # 原本字符串格式
        #             parsed_date = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S")

        # except Exception as e:
        #         print(f"[日期解析失败] {raw_date} -> {e}")
        #         parsed_date = None
        # print(f"[save orders to xlsx]Date parsed is {raw_date}")
        parsed_date=None
        try:
            parsed_date = datetime.strptime(raw_date, "%m/%d/%Y %H:%M")
        except Exception:
            pass
        existing_orders[order_id] = [

                store,  # 新增 store

                "'" + order.get('order_id', ''),
                parsed_date,
                order.get('buyer', ''),
                order.get('product', ''),
                order.get('specs', ''),
                order.get('sku', ''),
                order.get('price', ''),
                order.get('qty', ''),
                order.get('amount', ''),
                order.get('status', ''),
                order.get('status_en', ''),
                order.get('ae_ioss', ''),
                order.get('semi_managed', ''),
                order.get('action', ''),
                order.get('recipient', ''),
                order.get('address', ''),
                order.get('postal_code', ''),
                order.get('email', ''),
                order.get('phone', ''),
                order.get('tax_number', ''),
                order.get('order_link', ''),
            ]

    # ── 排序 ─────────────────────────────────────────
    def sort_key(row):
        date_val = row[2]
        if isinstance(date_val, datetime):
            return date_val
        return datetime.min

    sorted_rows = sorted(existing_orders.values(), key=sort_key, reverse=True)
    # ── 写入 Excel ──────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, values in enumerate(sorted_rows, 2):

        for col, value in enumerate(values, 1):

            cell = ws.cell(row=row_idx, column=col, value=value)

            if col == 4 and isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

        link_val = values[-1]

        if link_val:
            ws.cell(row=row_idx, column=len(values)).hyperlink = link_val
            ws.cell(row=row_idx, column=len(values)).font = Font(color="0563C1", underline="single")

    col_widths = [10,20,18,12,40,25,15,12,6,12,16,20,8,14,20,20,50,12,25,15,15, 75]

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(filepath)

    print(f"OK 已保存 {len(existing_orders)} 条订单（含历史） -> {filepath}")

    wb.close()

    time.sleep(0.5)

    return filepath, filename


